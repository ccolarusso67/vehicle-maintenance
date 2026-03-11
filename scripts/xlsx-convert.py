#!/usr/bin/env python3
"""Convert Ravenol XLSX files to CSV and merge with existing CSV data."""

import csv
import os
from openpyxl import load_workbook

BASE = '/Users/carminecolarusso/Desktop/Claude'
OUTPUT = os.path.join(BASE, 'Ravenol-combined.csv')

EXPECTED_HEADERS = ['makeId', 'makeLabel', 'modelId', 'modelLabel', 'typeId', 'typeLabel',
                    'partName', 'partCode', 'capacities', 'useType', 'intervals',
                    'productCode', 'productLabel']

def dedup_key(row):
    """Create dedup key from first 7 fields (make through partName)."""
    return '|'.join(str(row[i]) if i < len(row) else '' for i in range(7))

def read_csv_data(path):
    """Read CSV and return rows as lists."""
    print(f"Reading {os.path.basename(path)}...")
    rows = []
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.reader(f)
        header_found = False
        for i, row in enumerate(reader):
            if not header_found:
                if len(row) >= 7 and 'makeId' in row and 'makeLabel' in row:
                    header_found = True
                    print(f"  Header found at line {i+1}")
                continue
            if len(row) >= 7 and row[1]:  # Must have makeLabel
                rows.append(row[:13])  # Only first 13 columns
    print(f"  {len(rows)} data rows")
    return rows

def read_xlsx_data(path):
    """Read XLSX file using openpyxl read-only mode for memory efficiency."""
    print(f"\nReading {os.path.basename(path)} (streaming mode)...")
    wb = load_workbook(path, read_only=True, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        if 'statistic' in sheet_name.lower():
            print(f"  Skipping sheet '{sheet_name}'")
            continue

        print(f"  Processing sheet '{sheet_name}'...")
        ws = wb[sheet_name]
        header_found = False
        col_offset = 0  # Track if data starts at column B (offset=1)
        sheet_rows = 0

        for row in ws.iter_rows(values_only=True):
            if not header_found:
                str_row = [str(c) if c else '' for c in row]
                if 'makeId' in str_row and 'makeLabel' in str_row:
                    header_found = True
                    # Detect column offset: find where 'makeId' starts
                    col_offset = str_row.index('makeId')
                    print(f"    Header found (column offset: {col_offset})")
                continue

            # Convert to strings, applying column offset
            all_vals = [str(c) if c is not None else '' for c in row]
            str_row = all_vals[col_offset:col_offset + 13]
            if len(str_row) >= 7 and str_row[1]:  # Must have makeLabel
                all_rows.append(str_row)
                sheet_rows += 1

                if sheet_rows % 100000 == 0:
                    print(f"    ...{sheet_rows} rows processed")

        print(f"    {sheet_rows} data rows from this sheet")

    wb.close()
    return all_rows

def build_product_lookup(csv_rows):
    """Build productCode→productLabel lookup from CSV rows that have proper Ravenol names."""
    lookup = {}
    for row in csv_rows:
        if len(row) >= 13:
            code = row[11].strip()   # productCode
            label = row[12].strip()  # productLabel
            # Only store if label is a real Ravenol name (not numeric)
            if code and label and not label.isdigit() and label.startswith('RAVENOL'):
                lookup[code] = label
    print(f"  Built product lookup: {len(lookup)} code→name mappings")
    return lookup

def resolve_products(rows, lookup):
    """Replace numeric productLabel values with Ravenol names from lookup."""
    resolved = 0
    for row in rows:
        if len(row) >= 13:
            label = row[12].strip()
            code = row[11].strip()
            if label.isdigit() and code in lookup:
                row[12] = lookup[code]
                resolved += 1
    return resolved

def main():
    print("=== Merging all Ravenol data files ===\n")

    seen = set()
    unique_rows = []

    # 1. Read existing CSV (baseline)
    csv_rows = read_csv_data(os.path.join(BASE, 'Ravenol rev.csv'))

    # Build product lookup from CSV (has proper Ravenol names)
    product_lookup = build_product_lookup(csv_rows)

    for row in csv_rows:
        key = dedup_key(row)
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)
    print(f"  {len(unique_rows)} unique rows from CSV\n")

    # 2. Read Ravenol.xlsx
    xlsx1_rows = read_xlsx_data(os.path.join(BASE, 'Ravenol.xlsx'))
    resolved1 = resolve_products(xlsx1_rows, product_lookup)
    print(f"  Resolved {resolved1} numeric product labels in Ravenol.xlsx")
    added_xlsx1 = 0
    for row in xlsx1_rows:
        key = dedup_key(row)
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)
            added_xlsx1 += 1
    print(f"  {added_xlsx1} new unique rows from Ravenol.xlsx\n")

    # 3. Read Ravenol-2.xlsx
    xlsx2_rows = read_xlsx_data(os.path.join(BASE, 'Ravenol-2.xlsx'))
    resolved2 = resolve_products(xlsx2_rows, product_lookup)
    print(f"  Resolved {resolved2} numeric product labels in Ravenol-2.xlsx")
    added_xlsx2 = 0
    for row in xlsx2_rows:
        key = dedup_key(row)
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)
            added_xlsx2 += 1
    print(f"  {added_xlsx2} new unique rows from Ravenol-2.xlsx\n")

    # 4. Write combined CSV with the 15-line header block parse-csv.mjs expects
    print(f"Writing combined CSV...")
    with open(OUTPUT, 'w', encoding='utf-8', newline='') as f:
        # Write 14 blank lines, then header, then data
        for _ in range(14):
            f.write('\n')
        writer = csv.writer(f)
        writer.writerow(EXPECTED_HEADERS)
        writer.writerows(unique_rows)

    print(f"\n=== Summary ===")
    print(f"CSV rows:           {len(csv_rows)}")
    print(f"Ravenol.xlsx rows:  {len(xlsx1_rows)} ({added_xlsx1} new, {resolved1} products resolved)")
    print(f"Ravenol-2.xlsx rows: {len(xlsx2_rows)} ({added_xlsx2} new, {resolved2} products resolved)")
    print(f"Total unique rows:  {len(unique_rows)}")
    print(f"Output: {OUTPUT}")

if __name__ == '__main__':
    main()
